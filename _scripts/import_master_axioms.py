r"""Import the Master_Axiom.xlsx registry into canonical axiom atoms.

This script is intentionally read-only against the workbook and repeatable
against the repo. It generates:

  - _vocab/axiom_registry.json
  - axioms/01_canonical/AX-###-slug.jsonld

Usage:
  python _scripts/import_master_axioms.py
  python _scripts/import_master_axioms.py C:\path\to\Master_Axiom.xlsx
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = Path(
    r"C:\theophysics\_CANONICAL_STRIPPED_OUT_REVIEW_2026-07-26"
    r"\CANONICAL\03_AXIOMS\01_canonical\Master_Axiom.xlsx"
)
OUT_DIR = REPO / "axioms" / "01_canonical"
REGISTRY_PATH = REPO / "_vocab" / "axiom_registry.json"

TYPE_TO_CLAIM_CLASS = {
    "Primitive": "floor_axiom",
    "Definition": "definition",
    "Theorem": "theorem",
    "FrameworkCommitment": "floor_axiom",
    "Equation": "mathematical",
    "Property": "theorem",
    "EvidenceNode": "empirical_anchor",
    "BoundaryCondition": "boundary",
    "OpenProblem": "prediction",
    "Identification": "theological_interpretation",
    "Corollary": "theorem",
    "Hypothesis": "prediction",
    "Prediction": "prediction",
    "Protocol": "empirical",
    "FalsificationCriterion": "empirical_anchor",
    "UniversalPrinciple": "floor_axiom",
    "Operator": "mathematical",
    "ObservableDomain": "empirical_anchor",
    "CapstoneTerminalClaim": "theorem",
    "BridgePrinciple": "bridge",
    "MetaClaim": "theorem",
    "ClosureClaim": "theorem",
}

TYPE_STATUS = {
    "Primitive": "active",
    "Definition": "active",
    "Theorem": "proposed",
    "FrameworkCommitment": "proposed",
    "Equation": "active",
    "Property": "active",
    "EvidenceNode": "classified",
    "BoundaryCondition": "proposed",
    "OpenProblem": "captured",
    "Identification": "proposed",
    "Corollary": "proposed",
    "Hypothesis": "proposed",
    "Prediction": "proposed",
    "Protocol": "captured",
    "FalsificationCriterion": "classified",
    "UniversalPrinciple": "proposed",
    "Operator": "active",
    "ObservableDomain": "classified",
    "CapstoneTerminalClaim": "proposed",
    "BridgePrinciple": "proposed",
    "MetaClaim": "proposed",
    "ClosureClaim": "proposed",
}

STATUS_RANK = {
    "captured": 0,
    "classified": 1,
    "proposed": 2,
    "active": 3,
}
RANK_STATUS = {value: key for key, value in STATUS_RANK.items()}

TAG_KEYWORDS = {
    "grace": ["grace"],
    "sin": ["sin"],
    "justice": ["justice"],
    "mercy": ["mercy"],
    "cross": ["cross"],
    "resurrection": ["resurrection"],
    "faith": ["faith"],
    "logos": ["logos", "chi", "logoi"],
    "coherence": ["coherence", "coherent"],
    "decoherence": ["decoherence", "decoherent"],
    "kill-condition": ["falsification", "defeat", "kill"],
    "covenant": ["covenant"],
    "atonement": ["atonement"],
    "repentance": ["repentance"],
    "sanctification": ["sanctification"],
    "witness": ["witness", "observer", "observation"],
    "entropy": ["entropy"],
    "boundary": ["boundary"],
    "glory": ["glory", "heaven"],
    "invariant": ["invariant"],
    "definition": ["definition"],
    "variable": ["variable", "operator"],
    "equation": ["equation", "lagrangian", "friedmann"],
    "proof": ["proof", "theorem"],
    "isomorphism": ["isomorphism"],
    "free-will": ["free will", "voluntary"],
    "moral-conservation": ["moral", "sign conservation"],
    "observer": ["observer", "observation", "collapse"],
    "information": ["information", "bit"],
    "symmetry": ["symmetry"],
    "conservation": ["conservation", "conserved"],
    "phase-transition": ["phase transition"],
    "curvature": ["curvature", "geometry", "spacetime"],
    "field": ["field", "scalar"],
    "regress": ["regress"],
    "trinity": ["trinity", "three observers"],
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:80] or "untitled"


def split_dependencies(raw: str) -> list[str]:
    if not raw:
        return []

    deps: list[str] = []
    for token in re.split(r"[,;]", raw):
        token = token.strip()
        if not token or token.startswith("∅"):
            continue

        for start, end in re.findall(r"\bAX-(\d{3})\s*-\s*(\d{3})\b", token):
            for n in range(int(start), int(end) + 1):
                deps.append(f"AX-{n:03d}")
            token = re.sub(r"\bAX-\d{3}\s*-\s*\d{3}\b", "", token)

        for match in re.findall(r"\bAX-\d{3}[a-z]?\b", token):
            deps.append(match)

    seen = set()
    out = []
    for dep in deps:
        if dep not in seen:
            seen.add(dep)
            out.append(dep)
    return out


def tags_for(row: dict[str, str]) -> list[str]:
    haystack = " ".join(
        [row.get("title", ""), row.get("claim", ""), row.get("domain", ""), row.get("new_type", "")]
    ).lower()
    tags = []
    for tag, needles in TAG_KEYWORDS.items():
        if any(needle in haystack for needle in needles):
            tags.append(tag)
    return tags[:8]


def plain_statement(row: dict[str, str]) -> str:
    claim = row.get("claim", "")
    title = row.get("title", "")
    if claim:
        return claim
    if title:
        return title
    return row.get("axiom_id", "Untitled axiom")


def load_rows(workbook: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook, read_only=True, data_only=False)
    ws = wb["Master Axiom"]
    header = [clean(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {header[i]: clean(values[i]) if i < len(values) else "" for i in range(len(header))}
        if not record.get("axiom_id", "").startswith("AX-"):
            continue
        record["source_sheet"] = ws.title
        record["source_row"] = str(excel_row)
        rows.append(record)
    return rows


def assign_monotone_statuses(rows: list[dict[str, str]]) -> None:
    """Lower imported status so no row outranks its dependencies."""
    by_id = {row["axiom_id"]: row for row in rows}
    desired = {
        row["axiom_id"]: TYPE_STATUS.get(row.get("new_type", ""), "proposed")
        for row in rows
    }
    ranks = {axiom_id: STATUS_RANK[status] for axiom_id, status in desired.items()}

    changed = True
    while changed:
        changed = False
        for row in rows:
            axiom_id = row["axiom_id"]
            dep_ranks = [
                ranks[dep]
                for dep in split_dependencies(row.get("depends_on", ""))
                if dep in by_id and dep != axiom_id
            ]
            if dep_ranks:
                ceiling = min([ranks[axiom_id], *dep_ranks])
                if ceiling < ranks[axiom_id]:
                    ranks[axiom_id] = ceiling
                    changed = True

    for row in rows:
        row["_intended_status"] = desired[row["axiom_id"]]
        row["_import_status"] = RANK_STATUS[ranks[row["axiom_id"]]]


def make_atom(row: dict[str, str], id_to_file: dict[str, str], source_path: Path) -> dict[str, Any]:
    axiom_id = row["axiom_id"]
    title = row.get("title") or axiom_id
    new_type = row.get("new_type", "")
    claim_class = TYPE_TO_CLAIM_CLASS.get(new_type, "theorem")
    status = row.get("_import_status") or TYPE_STATUS.get(new_type, "proposed")
    deps = split_dependencies(row.get("depends_on", ""))

    edges = []
    for dep in deps:
        if dep == axiom_id:
            continue
        target_file = id_to_file.get(dep)
        target = f"tp:axioms/01/{dep}" if target_file else f"tp:axioms/01/{dep}"
        edges.append(
            {
                "type": "dependsOn",
                "target": target,
                "propagates": True,
                "note": "Imported from Master Axiom depends_on column.",
            }
        )

    atom = {
        "@context": [
            "https://schema.org",
            "https://faiththruphysics.com/vocab/context.jsonld",
        ],
        "@type": "Claim",
        "@id": f"https://faiththruphysics.com/claims/axioms/01/{axiom_id}",
        "nodeID": f"tp:axioms/01/{axiom_id}",
        "claimID": f"tp:AXIOMS/{row.get('old_id') or axiom_id}",
        "name": title,
        "author": [
            {"@type": "Person", "name": "David Lowe"},
            {"@type": "SoftwareApplication", "name": "GPT (OpenAI)", "tp:role": "ai-collaborator"},
            {"@type": "SoftwareApplication", "name": "Claude (Anthropic)", "tp:role": "ai-collaborator"},
        ],
        "aiContributionDeclared": True,
        "dateCreated": "2026-07-28",
        "dateModified": date.today().isoformat(),
        "version": "0.1.0",
        "nodeType": "claim",
        "domainType": "axioms",
        "stage": "01_canonical",
        "status": status,
        "claimClass": claim_class,
        "statementTechnical": row.get("claim") or title,
        "statementPlain": plain_statement(row),
        "mathematicalForm": row.get("formal_expression", ""),
        "axiomRoot": "https://faiththruphysics.com/claims/axioms/01/AX-001",
        "edges": edges,
        "tags": tags_for(row),
        "keywords": [k for k in [row.get("old_id"), row.get("tree_level"), row.get("domain"), new_type] if k],
        "verificationStatus": "registry-import",
        "kernelChecked": False,
        "challengeStatus": "unchallenged",
        "falsificationCondition": row.get("defeat_conditions") or row.get("theory_question_left_open", ""),
        "sourceReference": str(source_path),
        "sourceWorkbook": {
            "path": str(source_path),
            "sheet": row.get("source_sheet", "Master Axiom"),
            "row": int(row.get("source_row", "0") or 0),
        },
        "axiomRegistry": {
            "axiomID": axiom_id,
            "oldID": row.get("old_id", ""),
            "treeLevel": row.get("tree_level", ""),
            "treeBranch": row.get("tree_branch", ""),
            "questionType": row.get("question_type", ""),
            "domain": row.get("domain", ""),
            "tier": row.get("tier", ""),
            "newType": new_type,
            "logicalForce": row.get("logical_force", ""),
            "formalizableInLean": row.get("formalizable_in_lean", ""),
            "leanKind": row.get("lean_kind", ""),
            "riskLevel": row.get("risk_level", ""),
            "kernelRole": row.get("kernel_role", ""),
            "moduleID": row.get("module_id", ""),
            "moduleTitle": row.get("module_title", ""),
            "spineRole": row.get("spine_role", ""),
            "boundaryCondition": row.get("boundary_condition", ""),
            "fourModeRelevance": row.get("four_mode_relevance", ""),
            "worldviewsEliminated": row.get("worldviews_eliminated", ""),
            "propagationTest": row.get("propagation_test", ""),
            "intendedStatus": row.get("_intended_status", status),
            "importStatusReason": "Status lowered when needed to satisfy dependency monotonicity.",
        },
    }

    if not atom["mathematicalForm"]:
        atom.pop("mathematicalForm")
    if not atom["falsificationCondition"]:
        atom.pop("falsificationCondition")
    if not atom["tags"]:
        atom.pop("tags")

    return atom


def main() -> int:
    workbook = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook.exists():
        raise SystemExit(f"Workbook not found: {workbook}")

    rows = load_rows(workbook)
    assign_monotone_statuses(rows)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    id_to_file = {
        row["axiom_id"]: f"{row['axiom_id']}-{slugify(row.get('title', ''))}.jsonld"
        for row in rows
    }

    for old in OUT_DIR.glob("AX-*.jsonld"):
        old.unlink()

    registry = {
        "sourceWorkbook": str(workbook),
        "sourceSheet": "Master Axiom",
        "generatedAt": date.today().isoformat(),
        "totalRows": len(rows),
        "countsByNewType": dict(Counter(row.get("new_type", "") for row in rows)),
        "countsByImportStatus": dict(Counter(row.get("_import_status", "") for row in rows)),
        "countsByIntendedStatus": dict(Counter(row.get("_intended_status", "") for row in rows)),
        "oldAxiomIDCount": sum(1 for row in rows if re.match(r"^A\d+\.\d+$", row.get("old_id", ""))),
        "oldAxiomIDCountsByNewType": dict(
            Counter(
                row.get("new_type", "")
                for row in rows
                if re.match(r"^A\d+\.\d+$", row.get("old_id", ""))
            )
        ),
        "items": [],
    }

    for row in rows:
        filename = id_to_file[row["axiom_id"]]
        atom = make_atom(row, id_to_file, workbook)
        (OUT_DIR / filename).write_text(
            json.dumps(atom, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        registry["items"].append(
            {
                "axiomID": row["axiom_id"],
                "oldID": row.get("old_id", ""),
                "title": row.get("title", ""),
                "newType": row.get("new_type", ""),
                "intendedStatus": row.get("_intended_status", ""),
                "status": atom["status"],
                "claimClass": atom["claimClass"],
                "domain": row.get("domain", ""),
                "tier": row.get("tier", ""),
                "sourceRow": int(row.get("source_row", "0") or 0),
                "atomPath": f"axioms/01_canonical/{filename}",
                "dependsOn": split_dependencies(row.get("depends_on", "")),
                "riskLevel": row.get("risk_level", ""),
                "kernelRole": row.get("kernel_role", ""),
            }
        )

    REGISTRY_PATH.write_text(
        json.dumps(registry, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Imported {len(rows)} axiom registry rows from {workbook}")
    print(f"Wrote {len(rows)} atoms to {OUT_DIR}")
    print(f"Wrote registry to {REGISTRY_PATH}")
    print("Counts by new_type:")
    for key, value in sorted(registry["countsByNewType"].items()):
        print(f"  {key}: {value}")
    print(
        "Old A*.* axiom IDs:",
        registry["oldAxiomIDCount"],
        registry["oldAxiomIDCountsByNewType"],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
