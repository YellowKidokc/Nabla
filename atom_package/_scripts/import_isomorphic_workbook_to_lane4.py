#!/usr/bin/env python3
"""Import Isomorphic_Updated.xlsx registry rows into Lane 4 atoms."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

import lane4_ledger as ledger


SOURCE = Path(r"\\192.168.2.50\h_hp\Desktop\Master EXCEL\Isomorphic_Updated.xlsx")
LOCAL_COPY = ledger.ROOT / "_runtime" / "Isomorphic_Updated.source-copy.xlsx"
REPORT = ledger.LEDGER / "LANE4_ISOMORPHIC_WORKBOOK_IMPORT_REPORT.md"


def value(cell):
    return "" if cell is None else str(cell).strip()


def find_header(ws, required):
    required = set(required)
    for row in ws.iter_rows(values_only=True):
        vals = [value(v) for v in row]
        if required.issubset(set(vals)):
            return vals
    raise SystemExit(f"header not found in sheet {ws.title}: {required}")


def rows_by_header(ws, required):
    header = find_header(ws, required)
    start = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [value(v) for v in row]
        if vals[: len(header)] == header[: len(vals)] or set(required).issubset(set(vals)):
            start = idx + 1
            break
    if not start:
        return []
    out = []
    for row in ws.iter_rows(min_row=start, values_only=True):
        vals = [value(v) for v in row]
        if not any(vals):
            continue
        item = {header[i]: vals[i] if i < len(vals) else "" for i in range(len(header))}
        out.append(item)
    return out


def c_level_starter(raw_level: str, verdict: str) -> str:
    text = f"{raw_level} {verdict}".lower()
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    level = max(nums) if nums else 0
    if "identity" in text or level >= 4:
        return "C4_REVIEW_IDENTITY_CANDIDATE"
    if "isomorphism" in text or level >= 3:
        return "C3_REVIEW_HOMOMORPHISM_OR_ISOMORPHISM_CANDIDATE"
    if "correspondence" in text or level >= 2:
        return "C2_PARTIAL_CORRESPONDENCE"
    if "analogy" in text or level >= 1:
        return "C0_ANALOGY"
    return "C0_UNCLASSIFIED_ANALOGY"


def claim(row):
    iso_id = row["ISO ID"]
    return (
        f"{iso_id}: {row['Name']} maps {row['Domain A (Physics/Math)']} "
        f"to {row['Domain B (Theology)']}. Workbook verdict: {row['My Verdict']}. "
        f"Honest verdict: {row['Honest One-Line Verdict']}"
    )


def atom_from(row, eval_row=None, promo_row=None, lean_rows=None):
    iso_id = row["ISO ID"]
    lean_rows = lean_rows or []
    negative_guards = [
        "Workbook ISO level is not automatic canon promotion.",
        "Do not label as C5 formal isomorphism without explicit preservation proof.",
        "Do not use as Lean proof unless the cited Lean file and theorem compile in the current package context.",
    ]
    if row.get("Weakest Link"):
        negative_guards.append(f"Weakest link: {row['Weakest Link']}")
    if promo_row:
        for key in ("Blocker #1", "Blocker #2", "Blocker #3"):
            if promo_row.get(key):
                negative_guards.append(f"Promotion blocker: {promo_row[key]}")
    kill_conditions = [
        "Mapping fails structural-preservation review.",
        "A rival explanation preserves the same features with fewer assumptions.",
        "A claimed Lean theorem cannot be located or compiled in current context.",
    ]
    if eval_row:
        for key in ("Test A: Structural Preservation", "Test B: Non-Arbitrariness", "Test C: Constraint", "Test D: Bidirectional"):
            if eval_row.get(key, "").lower() in {"fail", "failed"}:
                kill_conditions.append(f"{key} fails in source evaluation.")
    if row.get("Weakest Link"):
        kill_conditions.append(f"Weakest link remains unresolved: {row['Weakest Link']}")

    return {
        "title": f"{iso_id} - {row['Name']}",
        "claim": claim(row),
        "domain": "isomorphic-events",
        "lane": "IsomorphicEvents",
        "source_claim_id": f"ISO:{iso_id}",
        "claim_class": "isomorphic_event",
        "mode_classification": c_level_starter(row.get("AAA-000 Level", ""), row.get("My Verdict", "")),
        "assumptions": [
            row.get("Domain A (Physics/Math)", "domain_a_unspecified") or "domain_a_unspecified",
            row.get("Domain B (Theology)", "domain_b_unspecified") or "domain_b_unspecified",
            "AAA classification workbook accepted as source classification, not proof",
        ],
        "definitions": [row.get("Strongest Feature", ""), row.get("Honest One-Line Verdict", "")],
        "equations": [],
        "bridges": [f"{row.get('Domain A (Physics/Math)', '')} -> {row.get('Domain B (Theology)', '')}"],
        "dependencies": [iso_id],
        "negative_guards": negative_guards,
        "kill_conditions": kill_conditions,
        "proof_label": "ISOMORPHIC_EVENT_CANDIDATE",
        "current_status": "active_candidate",
        "rerun_status": "not_applicable",
        "source_artifacts": [str(SOURCE), str(LOCAL_COPY)],
        "workbook_claimed_level": row.get("Claimed Level", ""),
        "workbook_aaa_level": row.get("AAA-000 Level", ""),
        "workbook_verdict": row.get("My Verdict", ""),
        "test_score": eval_row.get("Score", "") if eval_row else "",
        "test_verdict": eval_row.get("Verdict", "") if eval_row else "",
        "test_evidence": eval_row.get("Key Evidence / Reasoning", "") if eval_row else "",
        "promotion_target": promo_row.get("Target Level", "") if promo_row else "",
        "promotion_status": promo_row.get("Status", "") if promo_row else "",
        "lean_receipts": lean_rows,
    }


def main():
    if not LOCAL_COPY.is_file():
        if not SOURCE.is_file():
            raise SystemExit(f"source not found: {SOURCE}")
        LOCAL_COPY.parent.mkdir(parents=True, exist_ok=True)
        LOCAL_COPY.write_bytes(SOURCE.read_bytes())

    wb = load_workbook(LOCAL_COPY, read_only=True, data_only=True)
    registry = rows_by_header(wb["ISO Registry"], ["ISO ID", "Name", "My Verdict"])
    evals = {r["ISO ID"]: r for r in rows_by_header(wb["Evaluation Detail"], ["ISO ID", "Score", "Verdict"])}
    promos = {r["ISO ID"]: r for r in rows_by_header(wb["Promotion Tracker"], ["ISO ID", "Current Level", "Target Level"])}
    lean_map = {}
    for r in rows_by_header(wb["Lean Verification Map"], ["ISO ID", "Lean File", "Key Theorems"]):
        lean_map.setdefault(r["ISO ID"], []).append(r)

    created = skipped = 0
    by_c_level = {}
    for row in registry:
        iso_id = row.get("ISO ID")
        if not iso_id or iso_id == "ISO ID":
            continue
        atom = ledger.normalize_atom(atom_from(row, evals.get(iso_id), promos.get(iso_id), lean_map.get(iso_id)), LOCAL_COPY)
        by_c_level[atom["mode_classification"]] = by_c_level.get(atom["mode_classification"], 0) + 1
        if ledger.atom_path(atom["atom_id"]).exists():
            skipped += 1
            continue
        ledger.append_event(atom, {
            "event_type": "isomorphic_workbook_imported",
            "lane": "IsomorphicEvents",
            "result": "recorded",
            "artifact_path": str(SOURCE),
            "meaning": f"Imported {iso_id} from Isomorphic_Updated.xlsx as an isomorphic-event candidate.",
            "limits": "Workbook level is preserved as source classification, not automatic formal isomorphism or canon proof.",
            "reviewer": "Codex",
        })
        created += 1

    ledger.rebuild()
    failed = ledger.validate()

    lines = [
        "# Lane 4 Isomorphic Workbook Import Report",
        "",
        f"Source: `{SOURCE}`",
        f"Local copy: `{LOCAL_COPY}`",
        "",
        f"Registry rows parsed: **{len(registry)}**",
        f"Atoms created: **{created}**",
        f"Atoms skipped because already present: **{skipped}**",
        "",
        "## Starter C-Level Review Buckets",
        "",
    ]
    for key in sorted(by_c_level):
        lines.append(f"- `{key}`: {by_c_level[key]}")
    lines += [
        "",
        "## Guardrail",
        "",
        "```text",
        "These are isomorphic-event candidate atoms.",
        "Workbook Level 3 is not automatically C5 formal isomorphism.",
        "Lean references are preserved as receipts to check, not silently promoted.",
        "```",
        "",
        "## Validation",
        "",
        f"Lane 4 validation failed: `{failed}`",
        "",
    ]
    REPORT.write_text("\n".join(lines), encoding="utf-8")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
